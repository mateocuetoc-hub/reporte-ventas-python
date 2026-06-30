import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

archivo_entrada = "ventas_cliente.xlsx"
archivo_salida = "reporte_final.xlsx"

# Leer archivo Excel del cliente
df = pd.read_excel(archivo_entrada)

# Validar columnas necesarias
columnas_necesarias = ["producto", "cantidad", "precio_unitario"]

for columna in columnas_necesarias:
    if columna not in df.columns:
        print(f"Error: falta la columna '{columna}' en el archivo de entrada.")
        exit()

# Calcular total por venta
df["total"] = df["cantidad"] * df["precio_unitario"]

# Calcular resumen general
total_vendido = df["total"].sum()

resumen_productos = df.groupby("producto").agg({
    "cantidad": "sum",
    "total": "sum"
}).reset_index()

producto_mas_vendido = resumen_productos.loc[
    resumen_productos["cantidad"].idxmax(),
    "producto"
]

cantidad_mas_vendida = resumen_productos["cantidad"].max()

# Crear archivo Excel
with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Ventas", index=False)
    resumen_productos.to_excel(writer, sheet_name="Resumen", index=False)

# Aplicar formato con openpyxl
wb = load_workbook(archivo_salida)

for hoja in wb.sheetnames:
    ws = wb[hoja]

    # Formato de encabezados
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E78")
        celda.alignment = Alignment(horizontal="center")

    # Ajustar ancho de columnas
    for columna in ws.columns:
        largo_maximo = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value is not None:
                largo_maximo = max(largo_maximo, len(str(celda.value)))

        ws.column_dimensions[letra_columna].width = largo_maximo + 4

    # Formato moneda para columnas de precio y total
    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            if ws.cell(row=1, column=celda.column).value in ["precio_unitario", "total"]:
                celda.number_format = '$#,##0'

# Crear hoja de resumen ejecutivo
ws = wb.create_sheet("Resumen Ejecutivo")

ws["A1"] = "REPORTE DE VENTAS"
ws["A1"].font = Font(bold=True, size=16)

ws["A3"] = "Total vendido"
ws["B3"] = total_vendido
ws["B3"].number_format = '$#,##0'

ws["A4"] = "Producto más vendido"
ws["B4"] = producto_mas_vendido

ws["A5"] = "Cantidad vendida"
ws["B5"] = cantidad_mas_vendida

for celda in ["A3", "A4", "A5"]:
    ws[celda].font = Font(bold=True)

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 25

wb.save(archivo_salida)

print("Reporte generado correctamente.")
print(f"Total vendido: ${total_vendido}")
print(f"Producto más vendido: {producto_mas_vendido} ({cantidad_mas_vendida} unidades)")
print(f"Archivo creado: {archivo_salida}")