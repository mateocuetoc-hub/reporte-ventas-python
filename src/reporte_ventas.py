from pathlib import Path
import sys
import argparse
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

ARCHIVO_ENTRADA_DEFAULT = DATA_DIR / "ventas_cliente.xlsx"
ARCHIVO_SALIDA_DEFAULT = OUTPUT_DIR / "reporte_final.xlsx"

COLUMNAS_NECESARIAS = ["producto", "cantidad", "precio_unitario"]


def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    texto = texto.replace(" ", "_")
    texto = texto.replace("-", "_")

    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )

    return texto


def leer_archivo(ruta_entrada, separador_csv):
    extension = ruta_entrada.suffix.lower()

    if extension in [".xlsx", ".xls"]:
        return pd.read_excel(ruta_entrada)

    if extension == ".csv":
        return pd.read_csv(ruta_entrada, sep=separador_csv)

    print("Error: formato de archivo no soportado.")
    print("Formatos aceptados: .xlsx, .xls, .csv")
    print(f"Archivo recibido: {ruta_entrada}")
    sys.exit(1)


def validar_archivo(df):
    df.columns = [normalizar_texto(columna) for columna in df.columns]

    columnas_faltantes = []

    for columna in COLUMNAS_NECESARIAS:
        if columna not in df.columns:
            columnas_faltantes.append(columna)

    if len(columnas_faltantes) > 0:
        print("Error: faltan columnas en el archivo de entrada.")
        print("Columnas necesarias:")
        for columna in COLUMNAS_NECESARIAS:
            print(f"- {columna}")

        print("\nColumnas faltantes:")
        for columna in columnas_faltantes:
            print(f"- {columna}")

        print("\nColumnas encontradas:")
        for columna in df.columns:
            print(f"- {columna}")

        sys.exit(1)

    if df.empty:
        print("Error: el archivo de ventas está vacío.")
        sys.exit(1)

    df["producto"] = df["producto"].astype(str).str.strip()
    df["cantidad"] = pd.to_numeric(df["cantidad"], errors="coerce")
    df["precio_unitario"] = pd.to_numeric(df["precio_unitario"], errors="coerce")

    if df["producto"].eq("").any():
        print("Error: hay productos sin nombre.")
        sys.exit(1)

    if df["cantidad"].isnull().any():
        print("Error: hay valores inválidos en la columna 'cantidad'.")
        sys.exit(1)

    if df["precio_unitario"].isnull().any():
        print("Error: hay valores inválidos en la columna 'precio_unitario'.")
        sys.exit(1)

    if (df["cantidad"] <= 0).any():
        print("Error: la cantidad debe ser mayor a 0.")
        sys.exit(1)

    if (df["precio_unitario"] <= 0).any():
        print("Error: el precio unitario debe ser mayor a 0.")
        sys.exit(1)

    return df


def ajustar_columnas(ws):
    for columna in ws.columns:
        largo_maximo = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value is not None:
                largo_maximo = max(largo_maximo, len(str(celda.value)))

        ws.column_dimensions[letra_columna].width = largo_maximo + 4


def formatear_tabla(ws):
    color_encabezado = "1F4E78"
    borde_fino = Side(style="thin", color="D9E2F3")

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=color_encabezado)
        celda.alignment = Alignment(horizontal="center")
        celda.border = Border(bottom=borde_fino)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            encabezado = ws.cell(row=1, column=celda.column).value

            if encabezado in ["precio_unitario", "total"]:
                celda.number_format = '$#,##0'

            if encabezado == "cantidad":
                celda.number_format = '#,##0'

    ajustar_columnas(ws)


def crear_resumen_ejecutivo(
    wb,
    nombre_negocio,
    fecha_reporte,
    total_vendido,
    producto_mas_vendido,
    cantidad_mas_vendida
):
    if "Resumen Ejecutivo" in wb.sheetnames:
        del wb["Resumen Ejecutivo"]

    ws = wb.create_sheet("Resumen Ejecutivo", 0)

    ws["A1"] = "REPORTE DE VENTAS"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E78")

    ws["A2"] = nombre_negocio
    ws["A2"].font = Font(bold=True, size=13)

    ws["A4"] = "Fecha del reporte"
    ws["B4"] = fecha_reporte

    ws["A5"] = "Total vendido"
    ws["B5"] = total_vendido
    ws["B5"].number_format = '$#,##0'

    ws["A6"] = "Producto más vendido"
    ws["B6"] = producto_mas_vendido

    ws["A7"] = "Cantidad vendida"
    ws["B7"] = cantidad_mas_vendida

    for celda in ["A4", "A5", "A6", "A7"]:
        ws[celda].font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    ws_resumen = wb["Resumen"]

    grafico = BarChart()
    grafico.title = "Ventas por producto"
    grafico.y_axis.title = "Total vendido"
    grafico.x_axis.title = "Producto"

    datos = Reference(
        ws_resumen,
        min_col=3,
        min_row=1,
        max_row=ws_resumen.max_row
    )

    categorias = Reference(
        ws_resumen,
        min_col=1,
        min_row=2,
        max_row=ws_resumen.max_row
    )

    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.height = 8
    grafico.width = 16

    ws.add_chart(grafico, "D4")


def crear_hoja_metadatos(
    wb,
    nombre_negocio,
    fecha_reporte,
    ruta_entrada,
    ruta_salida,
    cantidad_registros,
    total_vendido
):
    if "Metadatos" in wb.sheetnames:
        del wb["Metadatos"]

    ws = wb.create_sheet("Metadatos")

    datos = [
        ["Nombre del negocio", nombre_negocio],
        ["Fecha del reporte", fecha_reporte],
        ["Archivo de entrada", str(ruta_entrada)],
        ["Archivo de salida", str(ruta_salida)],
        ["Cantidad de registros procesados", cantidad_registros],
        ["Total vendido", total_vendido],
        ["Herramienta", "Python + Pandas + OpenPyXL"],
        ["Versión del proyecto", "4"]
    ]

    for fila, dato in enumerate(datos, start=1):
        ws.cell(row=fila, column=1).value = dato[0]
        ws.cell(row=fila, column=2).value = dato[1]

    for celda in ws["A"]:
        celda.font = Font(bold=True)

    ws["B6"].number_format = '$#,##0'

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80


def generar_reporte(ruta_entrada, ruta_salida, nombre_negocio, separador_csv):
    ruta_entrada = Path(ruta_entrada)
    ruta_salida = Path(ruta_salida)

    if not ruta_entrada.is_absolute():
        ruta_entrada = BASE_DIR / ruta_entrada

    if not ruta_salida.is_absolute():
        ruta_salida = BASE_DIR / ruta_salida

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    if not ruta_entrada.exists():
        print("Error: no se encontró el archivo de entrada.")
        print(f"Ruta recibida: {ruta_entrada}")
        sys.exit(1)

    df = leer_archivo(ruta_entrada, separador_csv)
    df = validar_archivo(df)

    df["total"] = df["cantidad"] * df["precio_unitario"]

    resumen_productos = df.groupby("producto").agg({
        "cantidad": "sum",
        "total": "sum"
    }).reset_index()

    resumen_productos = resumen_productos.sort_values(
        by="total",
        ascending=False
    )

    total_vendido = df["total"].sum()

    producto_mas_vendido = resumen_productos.loc[
        resumen_productos["cantidad"].idxmax(),
        "producto"
    ]

    cantidad_mas_vendida = resumen_productos["cantidad"].max()

    fecha_reporte = datetime.now().strftime("%d-%m-%Y %H:%M")

    try:
        with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Ventas", index=False)
            resumen_productos.to_excel(writer, sheet_name="Resumen", index=False)

        wb = load_workbook(ruta_salida)

        formatear_tabla(wb["Ventas"])
        formatear_tabla(wb["Resumen"])

        crear_resumen_ejecutivo(
            wb,
            nombre_negocio,
            fecha_reporte,
            total_vendido,
            producto_mas_vendido,
            cantidad_mas_vendida
        )

        crear_hoja_metadatos(
            wb,
            nombre_negocio,
            fecha_reporte,
            ruta_entrada,
            ruta_salida,
            len(df),
            total_vendido
        )

        wb.save(ruta_salida)

    except PermissionError:
        print("Error: no se pudo guardar el reporte.")
        print("Cierra el archivo de salida si está abierto en LibreOffice.")
        sys.exit(1)

    print("Reporte generado correctamente.")
    print(f"Negocio: {nombre_negocio}")
    print(f"Fecha: {fecha_reporte}")
    print(f"Archivo de entrada: {ruta_entrada}")
    print(f"Archivo creado: {ruta_salida}")
    print(f"Total vendido: ${total_vendido:,.0f}")
    print(f"Producto más vendido: {producto_mas_vendido} ({cantidad_mas_vendida} unidades)")


def obtener_argumentos():
    parser = argparse.ArgumentParser(
        description="Generador automático de reportes de ventas en Excel."
    )

    parser.add_argument(
        "entrada",
        nargs="?",
        default=str(ARCHIVO_ENTRADA_DEFAULT),
        help="Ruta del archivo de entrada. Soporta .xlsx, .xls y .csv."
    )

    parser.add_argument(
        "-o",
        "--output",
        default=str(ARCHIVO_SALIDA_DEFAULT),
        help="Ruta del archivo Excel de salida."
    )

    parser.add_argument(
        "-n",
        "--negocio",
        default="Negocio de ejemplo",
        help="Nombre del negocio que aparecerá en el reporte."
    )

    parser.add_argument(
        "--sep",
        default=",",
        help="Separador para archivos CSV. Por defecto usa coma (,)."
    )

    return parser.parse_args()


if __name__ == "__main__":
    args = obtener_argumentos()

    generar_reporte(
        args.entrada,
        args.output,
        args.negocio,
        args.sep
    )